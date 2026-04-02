from __future__ import annotations

import csv
from contextlib import contextmanager
from pathlib import Path
from typing import Hashable, Iterable, Iterator, Mapping, Sequence

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

Row = dict[str, str]

from collections import Counter

def count_duplicates_sorted(
    items: Iterable[Hashable],
    threshold: int = 2,
    reverse: bool = True,
) -> list[tuple[Hashable, int]]:
    """Count items and return those occurring at least `threshold` times."""
    counter = Counter(items)
    duplicates = [(key, count) for key, count in counter.items() if count >= threshold]
    duplicates.sort(key=lambda x: x[1], reverse=reverse)
    return duplicates


@contextmanager
def _open_worksheet(excel_path: str | Path, *, sheet_name: str | None = None) -> Iterator[Worksheet]:
    """Yield a worksheet from an Excel workbook and close the workbook afterward."""
    workbook: Workbook = load_workbook(Path(excel_path), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        yield worksheet
    finally:
        workbook.close()


def iter_worksheet_rows(worksheet: Worksheet) -> Iterable[list[str]]:
    """Yield worksheet rows as CSV-ready strings."""
    for row in worksheet.iter_rows():
        yield ["" if cell.value is None else str(cell.value) for cell in row]


def write_worksheet_to_csv(worksheet: Worksheet, csv_path: str | Path) -> None:
    """Write worksheet rows to a CSV file."""
    csv_path = Path(csv_path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    with csv_path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerows(iter_worksheet_rows(worksheet))


def read_csv_rows(csv_path: Path | str, encoding: str = "utf-8") -> list[Row]:
    """Read a CSV file into a list of row dictionaries."""
    path = Path(csv_path)
    with path.open("r", encoding=encoding, newline="") as fp:
        reader = csv.DictReader(fp)
        return list(reader)


def write_csv_rows(
    csv_path: Path | str,
    rows: Sequence[Mapping[str, object]],
    *,
    fieldnames: Sequence[str],
    encoding: str = "utf-8",
) -> None:
    """Write row dictionaries to a CSV file."""
    path = Path(csv_path)
    with path.open("w", encoding=encoding, newline="") as fp:
        writer = csv.DictWriter(fp, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)
