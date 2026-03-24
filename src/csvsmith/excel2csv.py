from __future__ import annotations

import csv
from contextlib import contextmanager
from pathlib import Path
from typing import Iterable, Iterator

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


@contextmanager
def _open_worksheet(excel_path: str | Path, *, sheet_name: str | None = None) -> Iterator[Worksheet]:
    """Yield a worksheet from an Excel workbook and close the workbook afterward."""
    workbook: Workbook = load_workbook(Path(excel_path), read_only=True, data_only=True)
    try:
        worksheet = workbook[sheet_name] if sheet_name else workbook.active
        yield worksheet
    finally:
        workbook.close()


def iter_worksheet_rows(worksheet: Worksheet) -> Iterable[list[str]]:
    """Yield worksheet rows as CSV-ready strings."""
    for row in worksheet.iter_rows():
        yield ["" if cell.value is None else str(cell.value) for cell in row]


def write_worksheet_to_csv(worksheet: Worksheet, csv_path: str | Path) -> None:
    """Write worksheet rows to a CSV file."""
    csv_path = Path(csv_path)
    csv_path.parent.mkdir(parents=True, exist_ok=True)

    with csv_path.open("w", newline="", encoding="utf-8") as file:
        writer = csv.writer(file)
        writer.writerows(iter_worksheet_rows(worksheet))


def _resolve_csv_path(excel_path: Path, csv_path: str | Path | None) -> Path:
    """Return the output CSV path, defaulting to the Excel file's stem."""
    return Path(csv_path) if csv_path else excel_path.with_suffix(".csv")


def excel_to_csv(
    excel_path: str | Path,
    csv_path: str | Path | None = None,
    *,
    sheet_name: str | None = None,
) -> Path:
    """Convert one Excel worksheet into a CSV file."""
    excel_path = Path(excel_path)
    csv_path = _resolve_csv_path(excel_path, csv_path)

    with _open_worksheet(excel_path, sheet_name=sheet_name) as worksheet:
        write_worksheet_to_csv(worksheet, csv_path)

    return csv_path